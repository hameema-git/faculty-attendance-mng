from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from datetime import date
from .models import Attendance
from django.shortcuts import get_object_or_404
from django.http import HttpResponseNotAllowed
from django.db.models import Sum
import pandas as pd




def faculty_login(request):
    if request.method == "POST":
        # username = request.POST['username']
        # password = request.POST['password']
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)
        if user and hasattr(user, 'is_faculty') and user.is_faculty:
            login(request, user)
            return redirect('faculty_dashboard')
        else:
            return render(request, 'attendance/login.html', {'error': 'Invalid credentials or not faculty'})
    
    return render(request, "attendance/login.html")




from django.contrib.admin.views.decorators import staff_member_required


def admin_login(request):
    if request.method == "POST":
        # username = request.POST["username"]
        # password = request.POST["password"]
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)
        if user and hasattr(user, 'is_admin') and user.is_admin:
            login(request, user)
            return redirect("admin_dashboard")
        else:
            return render(request, "attendance/admin_login.html", {"error": "Invalid credentials or not an admin"})

    return render(request, "attendance/admin_login.html")



def home(request):
    return render(request, 'attendance/home.html')


from django.contrib.auth.decorators import login_required

@login_required
def admin_dashboard(request):
    if not request.user.is_admin:
        return redirect('home')  # or show access denied
    return render(request, 'attendance/admin_dashboard.html')

from django.contrib.auth import logout
from django.shortcuts import redirect
from django.shortcuts import render, redirect
from .models import CustomUser, Attendance
from attendance.forms import FacultyForm  # You need to define this form
from django.contrib.auth import logout

def custom_logout(request):
    logout(request)
    return redirect('home')  # Or 'faculty_login' or 'admin_login' if you prefer


# def admin_view_faculty(request):
#     faculty_list = CustomUser.objects.filter(is_faculty=True)
#     return render(request, 'attendance/admin_view_faculty.html', {'faculty_list': faculty_list})

from django.core.paginator import Paginator
from django.shortcuts import render
from .models import CustomUser


from django.core.paginator import Paginator

def admin_view_faculty(request):
    faculty_list = CustomUser.objects.filter(is_faculty=True).order_by('first_name')  # Optional: Order alphabetically
    paginator = Paginator(faculty_list, 10)  # Show 10 faculty per page

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'attendance/admin_view_faculty.html', {
        'page_obj': page_obj
    })



from datetime import date
from dateutil.relativedelta import relativedelta

from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from django.shortcuts import render

from .models import Attendance

def admin_view_attendance(request):
    # 1) Fetch filters from GET
    search_name  = request.GET.get('name', '').strip()
    search_date  = request.GET.get('date', '').strip()
    month_year   = request.GET.get('month_year', '').strip()

    selected_month = selected_year = None

    # 2) Define academic year range
    academic_start = date(2025, 5, 1)
    academic_end   = date(2026, 3, 31)

    # 3) Base queryset
    qs = Attendance.objects.filter(date__range=(academic_start, academic_end))

    # 4) Apply name filter
    if search_name:
        qs = qs.filter(
            Q(faculty__first_name__icontains=search_name) |
            Q(faculty__last_name__icontains=search_name)  |
            Q(faculty__username__icontains=search_name)
        )

    # 5) Date or Month-Year filter
    if search_date:
        qs = qs.filter(date=search_date)
    # elif month_year:
    #     try:
    #         m, y = map(int, month_year.split('-'))
    #         selected_month, selected_year = m, y
    #         qs = qs.filter(date__month=m, date__year=y)
    #     except ValueError:
    #         pass

    elif month_year:
        try:
            m, y = map(int, month_year.split('-'))
            if 1 <= m <= 12 and 1900 <= y <= 2100:
                selected_month, selected_year = m, y
                qs = qs.filter(date__month=m, date__year=y)
            else:
                # Optional: set a message or redirect with warning
                pass
        except ValueError:
            pass
    else:
        # default to current month (if within academic year)
        today = date.today()
        if academic_start <= today <= academic_end:
            selected_month, selected_year = today.month, today.year
            qs = qs.filter(date__month=selected_month, date__year=selected_year)

    # 6) Order by date descending
    qs = qs.order_by('-date')

    # 7) Pagination (10 records per page)
    paginator   = Paginator(qs, 9)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # 8) Build month/year dropdown options
    months = []
    current = academic_start
    while current <= academic_end:
        months.append({
            'month': current.month,
            'year': current.year,
            'label': current.strftime("%B %Y"),
            'value': f"{current.month}-{current.year}"
        })
        current += relativedelta(months=1)

    # 9) Render template with paginated page_obj
    return render(request, 'attendance/admin_view_attendance.html', {
        'page_obj':        page_obj,
        'search_name':     search_name,
        'search_date':     search_date,
        'months':          months,
        'selected_month':  selected_month,
        'selected_year':   selected_year,
    })


def admin_add_faculty(request):
    if request.method == 'POST':
        form = FacultyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_view_faculty')
    else:
        form = FacultyForm()
    return render(request, 'attendance/admin_add_faculty.html', {'form': form})

from django.shortcuts import render, redirect
from .models import Attendance,FacultyProfile

from django.core.paginator import Paginator
from django.shortcuts import render, redirect
from django.db.models import Q
from datetime import date
from dateutil.relativedelta import relativedelta
from .models import Attendance

def admin_approve_attendance(request):
    search_date = request.GET.get('date', '').strip()
    search_name = request.GET.get('name', '').strip()
    month_year = request.GET.get('month_year')
    selected_month = selected_year = None

    academic_start = date(2025, 5, 1)
    academic_end = date(2026, 3, 31)

    pending_attendance = Attendance.objects.filter(approved=False, date__range=(academic_start, academic_end))

    if search_name:
        pending_attendance = pending_attendance.filter(
            Q(faculty__first_name__icontains=search_name) |
            Q(faculty__last_name__icontains=search_name) |
            Q(faculty__username__icontains=search_name)
        )

    if search_date:
        pending_attendance = pending_attendance.filter(date=search_date)
    elif month_year:
        try:
            selected_month, selected_year = map(int, month_year.split('-'))
            pending_attendance = pending_attendance.filter(
                date__month=selected_month,
                date__year=selected_year
            )
        except ValueError:
            pass
    else:
        today = date.today()
        if academic_start <= today <= academic_end:
            selected_month = today.month
            selected_year = today.year
            pending_attendance = pending_attendance.filter(
                date__month=selected_month,
                date__year=selected_year
            )

    if request.method == 'POST':
        ids = request.POST.getlist('attendance_ids')
        Attendance.objects.filter(id__in=ids).update(approved=True)
        return redirect('admin_approve_attendance')

    # Pagination
    paginator = Paginator(pending_attendance.order_by('-date'), 5)  # Show 10 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Generate academic months
    months = []
    current = academic_start
    while current <= academic_end:
        months.append({
            'month': current.month,
            'year': current.year,
            'label': current.strftime("%B %Y"),
            'value': f"{current.month}-{current.year}"
        })
        current += relativedelta(months=1)

    return render(request, 'attendance/admin_approve_attendance.html', {
        'attendances': page_obj,
        'page_obj': page_obj,
        'search_name': search_name,
        'search_date': search_date,
        'months': months,
        'selected_month': selected_month,
        'selected_year': selected_year,
    })

@login_required
def faculty_dashboard(request):
    today = date.today()
    attendance_exists = Attendance.objects.filter(faculty=request.user, date=today).exists()
    return render(request, "attendance/faculty_dashboard.html", {
        'attendance_exists': attendance_exists,
        'today': today
    })


from datetime import date,datetime
from dateutil.relativedelta import relativedelta

from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import Attendance

@login_required
def faculty_attendance_history(request):
    month_year  = request.GET.get('month_year', '').strip()
    search_date = request.GET.get('date', '').strip()

    # Academic year range
    academic_start = date(2025, 5, 1)
    academic_end   = date(2026, 5, 31)

    # Base queryset for this user
    qs = Attendance.objects.filter(
        faculty=request.user,
        date__range=(academic_start, academic_end)
    )

    selected_month = selected_year = None

    # Exact date filter
    if search_date:
        try:
            d_obj = datetime.strptime(search_date, '%Y-%m-%d').date()
            qs = qs.filter(date=d_obj)
        except ValueError:
            pass

    # Month-year filter
    elif month_year:
        try:
            m, y = map(int, month_year.split('-'))
            selected_month, selected_year = m, y
            qs = qs.filter(date__month=m, date__year=y)
        except ValueError:
            pass

    # Default to current month
    else:
        today = date.today()
        if academic_start <= today <= academic_end:
            selected_month, selected_year = today.month, today.year
            qs = qs.filter(date__month=selected_month, date__year=selected_year)

    qs = qs.order_by('-date')

    # Paginate: 5 records per page
    paginator   = Paginator(qs, 7)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Build month list for dropdown
    months = []
    cur = academic_start
    while cur <= academic_end:
        months.append({
            'month': cur.month,
            'year': cur.year,
            'label': cur.strftime("%B %Y"),
            'value': f"{cur.month}-{cur.year}"
        })
        cur += relativedelta(months=1)

    return render(request, "attendance/faculty_history.html", {
        'page_obj':       page_obj,
        'months':         months,
        'selected_month': selected_month,
        'selected_year':  selected_year,
        'search_date':    search_date,
    })




@login_required
def admin_delete_faculty(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        user.delete()
        return redirect('admin_view_faculty')  # or use messages for confirmation

    return render(request, 'attendance/confirm_delete_faculty.html', {'user': user})



from .forms import EditFacultyForm,AttendanceForm,AttendanceForm1

def admin_edit_faculty(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    faculty_profile = get_object_or_404(FacultyProfile, user=user)

    if request.method == 'POST':
        form = EditFacultyForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            # Update FacultyProfile manually
            faculty_profile.department = form.cleaned_data['department']
            faculty_profile.joining_date = form.cleaned_data['joining_date']
            faculty_profile.personal_phone = form.cleaned_data['personal_phone']
            faculty_profile.guardian_phone = form.cleaned_data['guardian_phone']
            faculty_profile.save()
            return redirect('admin_view_faculty')
    else:
        # Set initial data for both User and FacultyProfile fields
        initial_data = {
            'department': faculty_profile.department,
            'joining_date': faculty_profile.joining_date,
            'personal_phone': faculty_profile.personal_phone,
            'guardian_phone': faculty_profile.guardian_phone,
        }
        form = EditFacultyForm(instance=user, initial=initial_data)

    return render(request, 'attendance/admin_edit_faculty.html', {'form': form})


from django.db.models import Count, Q
from .models import Attendance, CustomUser

from calendar import monthrange
from datetime import date


from calendar import monthrange
from datetime import date
from django.shortcuts import render
from .models import Attendance, CustomUser
from django.db.models import Q

from datetime import date
from django.db.models import Q
from django.shortcuts import render
from .models import Attendance, CustomUser

from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta  # You may need to install python-dateutil

from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from django.db.models import Q
from .models import Attendance, CustomUser



from django.core.paginator import Paginator

def attendance_summary(request):
    today = date.today()

    academic_start = date(2025, 5, 1)
    academic_end = date(2026, 3, 31)

    try:
        selected_month = int(request.GET.get('month', today.month))
        selected_year = int(request.GET.get('year', today.year))
    except ValueError:
        selected_month = today.month
        selected_year = today.year

    if not (academic_start <= date(selected_year, selected_month, 1) <= academic_end):
        selected_month = 5
        selected_year = 2025

    faculty_users = CustomUser.objects.filter(is_faculty=True)

    search_query = request.GET.get('q', '').strip()
    if search_query:
        faculty_users = faculty_users.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )

    summary = []
    # for faculty in faculty_users:
    #     present_days = Attendance.objects.filter(
    #         faculty=faculty,
    #         date__month=selected_month,
    #         date__year=selected_year,
    #         status='P',
    #         approved=True
    #     ).count()

    #     absent_days = Attendance.objects.filter(
    #         faculty=faculty,
    #         date__month=selected_month,
    #         date__year=selected_year,
    #         status='A',
    #         approved=True
    #     ).count()

    #     leave_days = Attendance.objects.filter(
    #         faculty=faculty,
    #         date__month=selected_month,
    #         date__year=selected_year,
    #         status='L',
    #         approved=True
    #     ).count()

    #     attendance_records = Attendance.objects.filter(
    #         faculty=faculty,
    #         date__month=selected_month,
    #         date__year=selected_year,
    #         status='P',
    #         approved=True
    #     ).values('date', 'hours_worked')

    #     total_hours = 0
    #     daily_hours = []
    #     for record in attendance_records:
    #         hours = record['hours_worked'] or 0
    #         total_hours += hours
    #         daily_hours.append({
    #             'date': record['date'],
    #             'hours_worked': hours
    #         })

    #     summary.append({
    #         'faculty': faculty,
    #         'present': present_days,
    #         'absent': absent_days,
    #         'leave': leave_days,
    #         'total_hours': total_hours,
    #         'daily_hours': daily_hours
    #     })
    for faculty in faculty_users:
        present_days = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='P',
            approved=True
        ).count()

        absent_days = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='A',
            approved=True
        ).count()

        leave_days = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='L',
            # approved=True
        ).count()

        # ✅ Fetch present day details (date + hours)
        present_records = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='P',
            approved=True
        ).values('date', 'hours_worked')

        # ✅ Fetch leave day details (date + reason)
        leave_records = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='L',
            # approved=True
        ).values('date', 'message')

        # ✅ Calculate total hours
        total_hours = sum([record['hours_worked'] or 0 for record in present_records])

        attendance_records = Attendance.objects.filter(
            faculty=faculty,
            date__month=selected_month,
            date__year=selected_year,
            status='P',
            approved=True
        ).values('date', 'hours_worked')

        total_hours = 0
        daily_hours = []
        for record in attendance_records:
            hours = record['hours_worked'] or 0
            total_hours += hours
            daily_hours.append({
                'date': record['date'],
                'hours_worked': hours
            })


        # ✅ Append data to summary list
        summary.append({
            'faculty': faculty,
            'present': present_days,
            'absent': absent_days,
            'leave': leave_days,
            'total_hours': total_hours,
            'present_days': list(present_records),
            'leave_days': list(leave_records),
            'daily_hours': daily_hours
        })


    # Pagination
    paginator = Paginator(summary, 10)# 10 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    months = []
    current = academic_start
    while current <= academic_end:
        months.append({
            'month': current.month,
            'year': current.year,
            'label': current.strftime("%B %Y"),
            'value': f"{current.month}-{current.year}"
        })
        current += relativedelta(months=1)

    return render(request, 'attendance/attendance_summary.html', {
        'page_obj': page_obj,
        'summary': page_obj.object_list,  # only the current page's data
        'month': datetime(selected_year, selected_month, 1).strftime("%B %Y"),
        'months': months,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'search_query': search_query,
    })

import io
import calendar
from datetime import date
from django.http import FileResponse
from django.db.models import Q, Sum
from django.db import models
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def attendance_summary_pdf(request):
    today = date.today()
    selected_month = int(request.GET.get('month', today.month))
    selected_year = int(request.GET.get('year', today.year))
    search_query = request.GET.get('q', '').strip()

    month_name = calendar.month_name[selected_month]  # ✅ Convert month number to name

    faculty_users = CustomUser.objects.filter(is_faculty=True)
    if search_query:
        faculty_users = faculty_users.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )

    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Title'], alignment=1, textColor=colors.teal, fontSize=18, spaceAfter=20
    )

    elements = []
    elements.append(Paragraph(f"Attendance Summary – {month_name} {selected_year}", title_style))
    elements.append(Spacer(1, 12))

    data = [["Sl No", "Faculty", "Present", "Absent", "Leave", "Total Hours Worked"]]
    for i, faculty in enumerate(faculty_users, start=1):
        present = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='P', approved=True
        ).count()
        absent = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='A', approved=True
        ).count()
        leave = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='L', approved=True
        ).count()
        total_hours = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, approved=True
        ).aggregate(total=models.Sum('hours_worked'))['total'] or 0

        data.append([
            str(i),
            f"{faculty.first_name} {faculty.last_name}",
            str(present),
            str(absent),
            str(leave),
            str(total_hours)
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.teal),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    elements.append(table)
    pdf.build(elements)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f"attendance_{month_name}_{selected_year}.pdf")

# import io
# import pandas as pd
# from django.http import FileResponse, HttpResponse
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import A4
# from reportlab.lib.units import inch

# def attendance_summary_pdf(request):
#     today = date.today()
#     selected_month = int(request.GET.get('month', today.month))
#     selected_year = int(request.GET.get('year', today.year))
#     search_query = request.GET.get('q', '').strip()

#     faculty_users = CustomUser.objects.filter(is_faculty=True)
#     if search_query:
#         faculty_users = faculty_users.filter(
#             Q(first_name__icontains=search_query) |
#             Q(last_name__icontains=search_query) |
#             Q(username__icontains=search_query)
#         )

#     buffer = io.BytesIO()
#     p = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4

#     p.setFont("Helvetica-Bold", 14)
#     p.drawCentredString(width / 2, height - inch, f"Attendance Summary - {selected_month}/{selected_year}")
#     p.setFont("Helvetica", 10)

#     y = height - 1.5 * inch
#     for faculty in faculty_users:
#         present = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='P').count()
#         absent = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='A').count()
#         leave = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='L').count()

#         p.drawString(1 * inch, y, f"{faculty.first_name} {faculty.last_name} | P:{present} | A:{absent} | L:{leave}")
#         y -= 0.3 * inch
#         if y < 1 * inch:
#             p.showPage()
#             y = height - 1 * inch

#     p.showPage()
#     p.save()

#     buffer.seek(0)
#     return FileResponse(buffer, as_attachment=True, filename=f"attendance_{selected_month}_{selected_year}.pdf")

import io
import pandas as pd
import calendar
from datetime import date
from django.http import HttpResponse
from django.db.models import Q, Sum
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def attendance_summary_excel(request):
    today = date.today()
    selected_month = int(request.GET.get('month', today.month))
    selected_year = int(request.GET.get('year', today.year))
    search_query = request.GET.get('q', '').strip()
    month_name = calendar.month_name[selected_month]  # ✅ Convert month number → name

    faculty_users = CustomUser.objects.filter(is_faculty=True)
    if search_query:
        faculty_users = faculty_users.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )

    data = []
    for faculty in faculty_users:
        present = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='P'
        ).count()
        absent = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='A'
        ).count()
        leave = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='L'
        ).count()
        total_hours = Attendance.objects.filter(
            faculty=faculty, date__month=selected_month, date__year=selected_year, status='P'
        ).aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0

        data.append({
            'Faculty Name': f"{faculty.first_name} {faculty.last_name}",
            'Username': faculty.username,
            'Days Present': present,
            'Days Absent': absent,
            'Leaves Taken': leave,
            'Total Hours Worked': total_hours,
        })

    # ✅ Create DataFrame
    df = pd.DataFrame(data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'{month_name}_{selected_year}')

        # ✅ Get worksheet for styling
        workbook = writer.book
        worksheet = writer.sheets[f'{month_name}_{selected_year}']

        # --- Styles ---
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ✅ Apply styles to header
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # ✅ Apply alignment & border to all cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

        # ✅ Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # column number
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # ✅ Add title row on top
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"Faculty Attendance Summary - {month_name} {selected_year}"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ✅ Prepare and return response
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_{month_name}_{selected_year}.xlsx'
    return response


# def attendance_summary_excel(request):
#     today = date.today()
#     selected_month = int(request.GET.get('month', today.month))
#     selected_year = int(request.GET.get('year', today.year))
#     search_query = request.GET.get('q', '').strip()

#     faculty_users = CustomUser.objects.filter(is_faculty=True)
#     if search_query:
#         faculty_users = faculty_users.filter(
#             Q(first_name__icontains=search_query) |
#             Q(last_name__icontains=search_query) |
#             Q(username__icontains=search_query)
#         )

#     data = []
#     for faculty in faculty_users:
#         present = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='P').count()
#         absent = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='A').count()
#         leave = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='L').count()
#         total_hours = Attendance.objects.filter(faculty=faculty, date__month=selected_month, date__year=selected_year, status='P').aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0

#         data.append({
#             'Faculty': f"{faculty.first_name} {faculty.last_name}",
#             'Username': faculty.username,
#             'Present': present,
#             'Absent': absent,
#             'Leave': leave,
#             'Total Hours Worked': total_hours,
#         })

#     df = pd.DataFrame(data)
#     buffer = io.BytesIO()
#     with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Attendance')

#     buffer.seek(0)
#     response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename=attendance_{selected_month}_{selected_year}.xlsx'
#     return response





from django.utils import timezone
from .models import Attendance
from django.contrib import messages
from decimal import Decimal, InvalidOperation



@login_required
def edit_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance updated successfully.")
            return redirect('admin_approve_attendance')
    else:
        form = AttendanceForm(instance=attendance)

    return render(request, 'attendance/edit_attendance.html', {'form': form, 'attendance': attendance})

@login_required
def edit_attendance1(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance updated successfully.")
            return redirect('admin_view_attendance')
    else:
        form = AttendanceForm(instance=attendance)

    return render(request, 'attendance/edit_attendance1.html', {'form': form, 'attendance': attendance})
@login_required
def view_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)

    return render(request, 'attendance/view_attendance.html', {
        'attendance': attendance
    })

@login_required
def edit_attendance2(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        form = AttendanceForm1(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance updated successfully.")
            return redirect('faculty_attendance_history')
    else:
        form = AttendanceForm(instance=attendance)

    return render(request, 'attendance/edit_attendance2.html', {'form': form, 'attendance': attendance})

@login_required
def delete_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, "Attendance deleted successfully.")
        return redirect('faculty_attendance_history')

    return render(request, 'attendance/confirm_delete_attendance.html', {'attendance': attendance})
@login_required
def delete_attendance1(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, "Attendance deleted successfully.")
        return redirect('admin_view_attendance')

    return render(request, 'attendance/confirm_delete_attendance1.html', {'attendance': attendance})

@login_required
def delete_attendance2(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, "Attendance deleted successfully.")
        return redirect('admin_approve_attendance')

    return render(request, 'attendance/confirm_delete_attendance2.html', {'attendance': attendance})

@login_required
def mark_attendance_page(request):
    today = date.today()
    attendance_exists = Attendance.objects.filter(faculty=request.user, date=today).exists()

    return render(request, 'attendance/mark_attendance_page.html', {
        'today': today,
        'attendance_exists': attendance_exists
    })

@login_required
def mark_leave_page(request):
    today = date.today()
    attendance_exists = Attendance.objects.filter(faculty=request.user, date=today).exists()
    return render(request, "attendance/mark_leave.html", {
        'attendance_exists': attendance_exists
    })

@login_required
def mark_leave(request):
    today = date.today()

    if Attendance.objects.filter(faculty=request.user, date=today).exists():
        messages.warning(request, "You have already marked attendance or leave today.")
        return redirect('mark_leave_page')
        # return redirect('mark_leave')


    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()

        Attendance.objects.create(
            faculty=request.user,
            date=today,
            status='L',
            message=reason
        )
        messages.success(request, "✅ Leave marked successfully for today.")
        return redirect('faculty_dashboard')

    return render(request, "attendance/mark_leave.html")



@login_required
def mark_attendance(request):
    if request.method == 'POST':
        today = date.today()

        if Attendance.objects.filter(faculty=request.user, date=today).exists():
            messages.warning(request, "You have already marked attendance for today.")
            return redirect('mark_attendance_page')

        hours_worked = request.POST.get('hours_worked')
        message = request.POST.get('message', '')

        Attendance.objects.create(
            faculty=request.user,
            date=today,
            status='P',
            hours_worked=hours_worked,
            message=message
        )
        messages.success(request, "✅ Attendance marked successfully!")
        return redirect('faculty_dashboard')

    # Optional: Instead of redirecting back and looping,
    # show a 405 Method Not Allowed
    return HttpResponseNotAllowed(['POST'])

def admin_view_faculty_detail(request, faculty_id):
    faculty = get_object_or_404(CustomUser, id=faculty_id, is_faculty=True)
    return render(request, 'attendance/admin_faculty_detail.html', {'faculty': faculty})


from django.shortcuts import get_object_or_404

def view_attendance_detail(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    return render(request, 'attendance/view_attendance_detail.html', {'attendance': attendance})

from django.core.management import call_command
from django.http import HttpResponse


def migrate_view(request):
    call_command('migrate')
    return HttpResponse("Migration completed.")



from django.http import HttpResponse
from django.contrib.auth import get_user_model

User = get_user_model()

def create_superuser(request):
    if User.objects.filter(username='admin').exists():
        return HttpResponse("Superuser already exists.")
    User.objects.create_superuser('admin', 'admin@example.com', 'admin123')
    return HttpResponse("Superuser created. Username: admin, Password: admin123")


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.management import call_command

@csrf_exempt
def run_auto_mark_absent(request):
    if request.method == "POST":
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if token == settings.SECRET_JOB_TOKEN:
            try:
                call_command("auto_mark_absent")
                return JsonResponse({"status": "success"})
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "unauthorized"}, status=403)




from django.http import JsonResponse
import os

@csrf_exempt
def run_job(request):
    token = request.headers.get("Authorization")
    expected_token = f"Bearer {os.getenv('SECRET_JOB_TOKEN', 'change-this-token')}"
    
    print("Received token:", token)
    print("Expected token:", expected_token)

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method", "method": request.method}, status=405)

    if token != expected_token:
        return JsonResponse({"error": "Unauthorized", "received_token": token}, status=401)

    from django.core.management import call_command
    call_command("auto_mark_absent")

    return JsonResponse({"status": "Job executed successfully"})


